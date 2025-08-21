import re
import os
import pandas as pd
import warnings
import dropbox
import webbrowser
import customtkinter as ctk
from datetime import timedelta
from dotenv import load_dotenv
from io import BytesIO
from openpyxl import load_workbook

warnings.simplefilter(action='ignore', category=pd.errors.SettingWithCopyWarning)
warnings.simplefilter("ignore", UserWarning)
load_dotenv(dotenv_path='misc/.env')
APP_KEY = os.getenv('DROPBOX_APP_KEY')
APP_SECRET = os.getenv('DROPBOX_APP_SECRET')

# def append_to_multiple_sheets(update_df, sheet_name, excel_path: str):
#     print("Saving all sheets")
#     sheet = excel_workbook[sheet_name]

#     # Append each DataFrame row to the sheet
#     for row in update_df.itertuples(index=False, name=None):
#         sheet.append(row)

#     # Save the workbook after processing all sheets
#     excel_workbook.save(excel_path)

def save_all_files():
    print("Saving all CSV files")
    global mvp_df, dnc_df, conv_df, db_id_df, pd_jr_aa_df
    final_df = pd.DataFrame(sum(outbound_df_list, []))
    final_df.drop_duplicates(inplace=True)

    final_seven_df = pd.DataFrame(sum(outbound_seven_days_list, []))
    final_seven_df.drop_duplicates(inplace=True)
    
    # with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     final_df.to_excel(writer, sheet_name='CallOut-14d+TextOut-30d', index=False, header=False)

    final_df.to_csv('./data/CallOut-14d+TextOut-30d (Cold).csv', index=False, header=None)
    final_seven_df.to_csv('./data/CallTextOut-7d (PD).csv', index=False, header=None)
    mvp_df.drop_duplicates().dropna().to_csv('./data/CCM+CH+MVPC+MVPT+JC+RC+PD (Cold).csv', index=False, header=None)
    dnc_df.drop_duplicates().dropna().to_csv('./data/DNC (Cold-PD).csv', index=False, header=None)
    conv_df.drop_duplicates().dropna().to_csv('./data/PDConvDup (PD).csv', index=False, header=None)
    db_id_df.drop_duplicates().dropna().to_csv('./data/UniqueDB ID (Cold).csv', index=False, header=None)
    pd_jr_aa_df.drop_duplicates().dropna().to_csv('./data/PDJRAADups (PD).csv', index=False, header=None)

def get_update_df(update_df: pd.DataFrame, sheet_name: str):
    pass
    # sheet = excel_workbook[sheet_name]

    # # Append each DataFrame row to the sheet
    # for row in update_df.itertuples(index=False, name=None):
    #     sheet.append(row)

    # # Save the workbook after processing all sheets
    # excel_workbook.save('./data/List Cleaner.xlsx')

def dropbox_authentication() -> str:
    auth_flow = dropbox.DropboxOAuth2FlowNoRedirect(APP_KEY, APP_SECRET)
    authorize_url = auth_flow.start()

    if authorize_url:
        webbrowser.open(authorize_url)

def export_to_dropbox(root_path: str, local_data_path: str, dbx: dropbox.Dropbox) -> None:

    print("Uploading to List Cleaner Files to Dropbox")
    sheet_names = [
        "CCM+CH+MVPC+MVPT+JC+RC+PD (Cold)",
        "DNC (Cold-PD)",
        "UniqueDB ID (Cold)",
        "CallOut-14d+TextOut-30d (Cold)",
        "CallTextOut-7d (PD)",
        "PDConvDup (PD)",
        "PDJRAADups (PD)"
    ]

    for sheet_name in sheet_names:
        dropbox_path = f"{root_path}/{sheet_name}.csv"
        local_file_path = f"{local_data_path}/{sheet_name}.csv"

        with open(local_file_path, 'rb') as f:
            dbx.files_upload(f.read(), dropbox_path, mode=dropbox.files.WriteMode.overwrite)

    print("Sucessfully uploaded Updated List Cleaner Files to Dropbox")

def read_dropbox_file(path: str, dbx):
    metadata, response = dbx.files_download(path)
    if path.endswith('.csv'):
        return pd.read_csv(BytesIO(response.content), low_memory=False, encoding_errors='replace')
    elif path.endswith('.xlsx'):
        return pd.read_excel(BytesIO(response.content))
    else:
        raise ValueError("Invalid file format: Please provide a .csv, .xlsx or .xlsb file.")
    
def add_pd_phones(path: str, dbx):
    print("Processing Pipedrive Phones Export")
    global mvp_df
    df = read_dropbox_file(path, dbx)
    df.drop(columns=['Deal - ID', 'Deal - Last RVM Date', 'Deal - RVM Dates'],
            axis=1,
            inplace=True)

    # Select only the phone columns (Phone 1 to Phone 10)
    phone_columns = [f'Person - Phone {i}' for i in range(1, 11)]
    phones_df = df[phone_columns]

    # Initialize a list to hold all phone numbers
    all_phone_numbers = []

    # Iterate over each column and each row to extract phone numbers
    for column in phone_columns:
        phones_df[column] = phones_df[column].fillna('')  # Replace NaN with an empty string
        for phone_entry in phones_df[column]:
            # Convert each entry to a string, handle floats (remove .0), and split comma-separated values
            phone_entry = str(phone_entry).replace('.0', '')  # Remove `.0` from floats
            valid_phone_pattern = re.compile(r'^\+?\d{10}$')  # A simple pattern for phone numbers (you can adjust it)
            all_phone_numbers.extend([phone.strip() for phone in phone_entry.split(',') if phone.strip() and valid_phone_pattern.match(phone.strip())])
            # all_phone_numbers.extend([phone.strip() for phone in phone_entry.split(',') if phone.strip()])

    # Create a DataFrame from the extracted phone numbers
    result_df = pd.DataFrame({'Phone Number': all_phone_numbers})

    result_df['Phone Number'] = result_df['Phone Number'] \
        .str.replace("(", "") \
        .str.replace(")", "") \
        .str.replace("-", "") \
        .str.replace(" ", "")
    
    result_df.drop_duplicates(subset=['Phone Number'], inplace=True)

    # get_update_df(result_df, 'CCM+CH+MVPC+MVPT+JC+RC+PD')
    result_df.columns = [0]
    mvp_df = pd.concat([mvp_df, result_df], ignore_index=True)
    result_df = None
    
    # # Use ExcelWriter to append the DataFrame to the existing file
    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        
    #     result_df.to_excel(writer, sheet_name='ContMgt+MVP+JC+PD+RC', index=False)

def add_unique_db(path: str, dbx):
    print("Processing Unique Database ID")
    global db_id_df
    df = read_dropbox_file(path, dbx)
    # Now handle the "Deal - Unique Database ID" column similarly
    deal_column = 'Deal - Unique Database ID'

    # Ensure "Deal - Unique Database ID" has NaN values replaced with empty strings
    df[deal_column] = df[deal_column].fillna('')

    # Initialize a list to hold all database IDs
    all_deals = []

    # Iterate over each row in the "Deal - Unique Database ID" column
    for deal_entry in df[deal_column]:
        # Split by " | " and add each entry to the list
        all_deals.extend([deal.strip() for deal in deal_entry.split('|') if deal.strip()])

    # Create a DataFrame for the Deal IDs
    deal_df = pd.DataFrame({'Deal - Unique Database ID': all_deals})
    deal_df['Deal - Unique Database ID'] = deal_df['Deal - Unique Database ID'].apply(lambda x: round(int(float(x))) if pd.notna(x) else x)
    deal_df['Deal - Unique Database ID'] = deal_df['Deal - Unique Database ID'].astype('Int64')
    deal_df.drop_duplicates(subset=['Deal - Unique Database ID'], inplace=True)

    # get_update_df(deal_df, 'UniqueDB ID')
    deal_df.columns = [0]
    db_id_df = pd.concat([db_id_df, deal_df], ignore_index=True)
    deal_df = None

    # # Use ExcelWriter to append the DataFrame to the existing file
    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        
    #     deal_df.to_excel(writer, sheet_name='UniqueDB ID', index=False)

def add_pd_jr_aa_dup(path: str, dbx):
    print("Processing Pipedrive JR AA Duplicate")
    global pd_jr_aa_df
    df = read_dropbox_file(path, dbx)
    phone_columns = [f'Person - Phone {i}' for i in range(2, 11)]
    phone_columns.extend(['Person - Phone - Work', 'Person - Phone - Home', 'Person - Phone - Mobile', 'Person - Phone - Other', 'Person - Archive - Phone'])
    phones_df = df[phone_columns]

    all_phone_numbers = []

    # Iterate over each column and each row to extract phone numbers
    for column in phone_columns:
        phones_df[column] = phones_df[column].fillna('')  # Replace NaN with an empty string
        for phone_entry in phones_df[column]:
            # Convert each entry to a string, handle floats (remove .0), and split comma-separated values
            phone_entry = str(phone_entry).replace('.0', '')  # Remove `.0` from floats
            valid_phone_pattern = re.compile(r'^\+?\d{10}$')  # A simple pattern for phone numbers (you can adjust it)
            all_phone_numbers.extend([phone.strip() for phone in phone_entry.split(',') if phone.strip() and valid_phone_pattern.match(phone.strip())])
            # all_phone_numbers.extend([phone.strip() for phone in phone_entry.split(',') if phone.strip()])

    # Create a DataFrame from the extracted phone numbers
    result_df = pd.DataFrame({'Phone Number': all_phone_numbers})

    result_df['Phone Number'] = result_df['Phone Number'] \
        .str.replace("(", "") \
        .str.replace(")", "") \
        .str.replace("-", "") \
        .str.replace(" ", "")
    
    result_df.drop_duplicates(subset=['Phone Number'], inplace=True)

    result_df.columns = [0]
    pd_jr_aa_df = result_df
    result_df = None

def add_pd_conv_dup(path: str, dbx):
    print("Processing Pipedrive Conversion Duplicate")
    global conv_df
    df = read_dropbox_file(path, dbx)
    phone_columns = [f'Person - Phone {i}' for i in range(2, 11)]
    phone_columns.extend(['Person - Phone - Work', 'Person - Phone - Home', 'Person - Phone - Mobile', 'Person - Phone - Other', 'Person - Archive - Phone'])
    phones_df = df[phone_columns]

    all_phone_numbers = []

    # Iterate over each column and each row to extract phone numbers
    for column in phone_columns:
        phones_df[column] = phones_df[column].fillna('')  # Replace NaN with an empty string
        for phone_entry in phones_df[column]:
            # Convert each entry to a string, handle floats (remove .0), and split comma-separated values
            phone_entry = str(phone_entry).replace('.0', '')  # Remove `.0` from floats
            valid_phone_pattern = re.compile(r'^\+?\d{10}$')  # A simple pattern for phone numbers (you can adjust it)
            all_phone_numbers.extend([phone.strip() for phone in phone_entry.split(',') if phone.strip() and valid_phone_pattern.match(phone.strip())])
            # all_phone_numbers.extend([phone.strip() for phone in phone_entry.split(',') if phone.strip()])

    # Create a DataFrame from the extracted phone numbers
    result_df = pd.DataFrame({'Phone Number': all_phone_numbers})

    result_df['Phone Number'] = result_df['Phone Number'] \
        .str.replace("(", "") \
        .str.replace(")", "") \
        .str.replace("-", "") \
        .str.replace(" ", "")
    
    result_df.drop_duplicates(subset=['Phone Number'], inplace=True)

    # get_update_df(result_df, 'PDConvDup')
    result_df.columns = [0]
    conv_df = pd.concat([conv_df, result_df], ignore_index=True)
    result_df = None

def add_remove_list(path: str, dbx):
    print("Processing Pipedrive Remove From List")
    global dnc_df
    df = read_dropbox_file(path, dbx)
    phone_columns = [f'Person - Phone {i}' for i in range(2, 11)]
    phone_columns.extend(['Person - Phone - Work', 'Person - Phone - Home', 'Person - Phone - Mobile', 'Person - Phone - Other', 'Person - Archive - Phone'])
    phones_df = df[phone_columns]

    all_phone_numbers = []

    # Iterate over each column and each row to extract phone numbers
    for column in phone_columns:
        phones_df[column] = phones_df[column].fillna('')  # Replace NaN with an empty string
        for phone_entry in phones_df[column]:
            # Convert each entry to a string, handle floats (remove .0), and split comma-separated values
            phone_entry = str(phone_entry).replace('.0', '')  # Remove `.0` from floats
            valid_phone_pattern = re.compile(r'^\+?\d{10}$')  # A simple pattern for phone numbers (you can adjust it)
            all_phone_numbers.extend([phone.strip() for phone in phone_entry.split(',') if phone.strip() and valid_phone_pattern.match(phone.strip())])
            # all_phone_numbers.extend([phone.strip() for phone in phone_entry.split(',') if phone.strip()])

    # Create a DataFrame from the extracted phone numbers
    result_df = pd.DataFrame({'Phone Number': all_phone_numbers})

    result_df['Phone Number'] = result_df['Phone Number'] \
        .str.replace("(", "") \
        .str.replace(")", "") \
        .str.replace("-", "") \
        .str.replace(" ", "")
    
    result_df.drop_duplicates(subset=['Phone Number'], inplace=True)

    # get_update_df(result_df, 'DNC')
    result_df.columns = [0]
    dnc_df = pd.concat([dnc_df, result_df], ignore_index=True)
    result_df = None

    # book = load_workbook(excel_file)

    # # Check if the sheet exists and load it into a DataFrame
    # if 'ContMgt+MVP+JC+PD+RC' in book.sheetnames:
    #     # Read the existing sheet into a DataFrame
    #     existing_df = pd.read_excel(excel_file, sheet_name='ContMgt+MVP+JC+PD+RC')
        
    #     # Append the new rows to the existing DataFrame
    #     updated_df = pd.concat([existing_df, result_df], ignore_index=True)
    # else:
    #     # If the sheet doesn't exist, just use the new DataFrame
    #     updated_df = result_df

    # # Write the updated DataFrame back to the sheet
    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        
    #     updated_df.to_excel(writer, sheet_name='ContMgt+MVP+JC+PD+RC', index=False, header=False)

def add_jc(path: str, dbx):
    print("Processing Just Call")
    global mvp_df
    metadata, response = dbx.files_download(path)
    df = pd.read_excel(BytesIO(response.content),
                       sheet_name="Messages Details",
                       header=6,
                       usecols=['Client Number', 'Delivery Status', 'Datetime'],
                       parse_dates=['Datetime'])
    thirty_days_ago = pd.to_datetime('today', utc=True, format='mixed', dayfirst=False) - timedelta(days=30)
    seven_days_ago = pd.to_datetime('today', utc=True, format='mixed', dayfirst=False) - timedelta(days=7)
    df['Datetime'] = pd.to_datetime(df['Datetime'], format='mixed', dayfirst=False, utc=True)
    df['Client Number'] = df['Client Number'].astype('Int64')
    result_df = df[df['Client Number'].notna()].copy()  # Ensure we work on a copy of the filtered DataFrame
    result_df['Client Number'] = result_df['Client Number'].astype('Int64').astype(str).str[1:]

    received_df = result_df[result_df['Delivery Status'].str.lower() == 'received'][['Client Number']]
    sent_df = result_df[(result_df['Delivery Status'].str.lower().isin(['sent', 'delivered'])) & (result_df['Datetime'] >= thirty_days_ago)][['Client Number']]
    sent_seven_df = result_df[(result_df['Delivery Status'].str.lower().isin(['sent', 'delivered'])) & (result_df['Datetime'] >= seven_days_ago)][['Client Number']]

    received_df.drop_duplicates(subset=['Client Number'], inplace=True)
    sent_df.drop_duplicates(subset=['Client Number'], inplace=True)
    sent_seven_df.drop_duplicates(subset=['Client Number'], inplace=True)

    # get_update_df(received_df, 'CCM+CH+MVPC+MVPT+JC+RC+PD')
    received_df.columns = [0]
    mvp_df = pd.concat([mvp_df, received_df], ignore_index=True)

    received_df = None
    # get_update_df(sent_df, 'JCSMS-Sent')

    # # Use ExcelWriter to append the DataFrame to the existing file
    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        
    #     received_df.to_excel(writer, sheet_name='JCSMS-Received', index=False, header=False)

    # with pd.ExcelWriter('./data/List Cleaner.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     sent_df.to_excel(writer, sheet_name='JCSMS-Sent', index=False, header=False)
    outbound_df_list.append(sent_df.values.tolist())
    sent_df = None

    outbound_seven_days_list.append(sent_seven_df.values.tolist())
    sent_seven_df = None

def add_sly(path: str, dbx):
    print("Processing Sly")
    global dnc_df
    df = read_dropbox_file(path, dbx)
    df.drop_duplicates(inplace=True)

    # get_update_df(df, 'DNC')
    df.columns = [0]
    dnc_df = pd.concat([dnc_df, df], ignore_index=True)

    df = None

    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        
    #     df.to_excel(writer, sheet_name='DNC', index=False, header=False)

def add_contact_center(df: pd.DataFrame, dbx):
    print("Processing Contact Center")
    global mvp_df
    fourteen_days_ago = pd.to_datetime('today', format='mixed', dayfirst=False) - timedelta(days=14)
    seven_days_ago = pd.to_datetime('today', format='mixed', dayfirst=False) - timedelta(days=7)
    df['Date'] = pd.to_datetime(df['Date'], format='mixed', dayfirst=False)
    df = df[df['Media Type Name'] != 'E-Mail']
    inbound_df = df[df['Skill Direction'] == 'Inbound'][['ANI/From']]
    outbound_df = df[(df['Skill Direction'] == 'Outbound') & (df['Date'] >= fourteen_days_ago)][['DNIS/To']]
    outbound_seven_df = df[(df['Skill Direction'] == 'Outbound') & (df['Date'] >= seven_days_ago)][['DNIS/To']]

    inbound_df.drop_duplicates(inplace=True)
    outbound_df.drop_duplicates(inplace=True)
    outbound_seven_df.drop_duplicates(inplace=True)

    # get_update_df(inbound_df, 'CCM+CH+MVPC+MVPT+JC+RC+PD')
    inbound_df.columns = [0]
    mvp_df = pd.concat([mvp_df, inbound_df], ignore_index=True)
    inbound_df = None
    # get_update_df(outbound_df, 'Outbound-2weeks')

    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     inbound_df.to_excel(writer, sheet_name='ContactMgtLogs', index=False, header=False)

    # with pd.ExcelWriter('./data/List Cleaner.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     outbound_df.to_excel(writer, sheet_name='Outbound-2weeks', index=False, header=False)

    outbound_df_list.append(outbound_df.values.tolist())
    outbound_df = None

    outbound_seven_days_list.append(outbound_seven_df.values.tolist())
    outbound_seven_df = None

def add_contact_history_inbound(df: pd.DataFrame, dbx):
    print("Processing Contact History")
    fourteen_days_ago = pd.to_datetime('today', format='mixed', dayfirst=False) - timedelta(days=14)
    seven_days_ago = pd.to_datetime('today', format='mixed', dayfirst=False) - timedelta(days=7)
    df['Start Time:'] = pd.to_datetime(df['Start Time:'], format='mixed', dayfirst=False)
    inbound_df = df[(df['Media Type'] != 'Email') & (df['Outbound'] == 0)][['ANI/From']]
    outbound_df = df[(df['Media Type'] != 'Email') & (df['Outbound'] == 1) & (df['Start Time:'] >= fourteen_days_ago)][['ANI/From']]
    outbound_seven_df = df[(df['Media Type'] != 'Email') & (df['Outbound'] == 1) & (df['Start Time:'] >= seven_days_ago)][['ANI/From']]

    inbound_df.drop_duplicates(inplace=True)
    outbound_df.drop_duplicates(inplace=True)
    outbound_seven_df.drop_duplicates(inplace=True)

    # get_update_df(inbound_df, 'CCM+CH+MVPC+MVPT+JC+RC+PD')
    global mvp_df
    inbound_df.columns = [0]
    mvp_df = pd.concat([mvp_df, inbound_df], ignore_index=True)
    inbound_df = None
    # get_update_df(outbound_df, 'CallOut-14d+TextOut-30d')
    outbound_df_list.append(outbound_df.values.tolist())
    outbound_df = None

    outbound_seven_days_list.append(outbound_seven_df.values.tolist())
    outbound_seven_df = None

def add_mvp_calls_inbound(df: pd.DataFrame, dbx):
    print("Processing MVP Calls")
    inbound_df = df[df['Call Direction'] == 'Inbound'][['From Number']]
    outbound_df = df[df['Call Direction'] == 'Outbound'][['From Number']]

    inbound_df['From Number'] = inbound_df['From Number'].str.replace(r'\D', '', regex=True)
    outbound_df['From Number'] = outbound_df['From Number'].str.replace(r'\D', '', regex=True)

    inbound_df.drop_duplicates(inplace=True)
    outbound_df.drop_duplicates(inplace=True)

    # get_update_df(inbound_df, 'CCM+CH+MVPC+MVPT+JC+RC+PD')
    global mvp_df
    inbound_df.columns = [0]
    mvp_df = pd.concat([mvp_df, inbound_df], ignore_index=True)
    inbound_df = None
    # get_update_df(outbound_df, 'CallOut-14d+TextOut-30d')
    outbound_df_list.append(outbound_df.values.tolist())
    outbound_df = None

def add_mvp(df: pd.DataFrame, dbx):
    print("Processing MVP Texts")
    # Calculate the date threshold
    thirty_days_ago = pd.Timestamp.utcnow() - timedelta(days=30)
    seven_days_ago = pd.Timestamp.utcnow() - timedelta(days=7)

    # Convert 'Date / Time' to datetime efficiently
    df['Date / Time'] = pd.to_datetime(df['Date / Time'], utc=True, errors='coerce')

    # Filter only valid numbers before conversion
    df = df[df['Sender Number'].notna()].copy()

    # Convert 'Sender Number' to string and slice efficiently
    df['Sender Number'] = df['Sender Number'].astype(str).str[1:]

    # Filter inbound and outbound numbers efficiently
    inbound_df = df.loc[
        (df['Direction'] == 'Inbound') & 
        (df['Sender Number'].str.len() == 10),
        ['Sender Number']
    ].drop_duplicates()

    outbound_df = df.loc[
        (df['Direction'] == 'Outbound') & 
        (df['Date / Time'] >= thirty_days_ago) & 
        (df['Sender Number'].str.len() == 10),
        ['Sender Number']
    ].drop_duplicates()

    outbound_seven_df = df.loc[
        (df['Direction'] == 'Outbound') & 
        (df['Date / Time'] >= seven_days_ago) & 
        (df['Sender Number'].str.len() == 10),
        ['Sender Number']
    ].drop_duplicates()

    # get_update_df(inbound_df, 'CCM+CH+MVPC+MVPT+JC+RC+PD')
    global mvp_df
    inbound_df.columns = [0]
    mvp_df = pd.concat([mvp_df, inbound_df], ignore_index=True)
    inbound_df = None
    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     final_df.to_excel(writer, sheet_name='MVPLogs', index=False, header=False)

    outbound_df_list.append(outbound_df.values.tolist())
    outbound_df = None

    outbound_seven_days_list.append(outbound_seven_df.values.tolist())
    outbound_seven_df = None

def add_rc(df: pd.DataFrame, dbx):
    print("Processing RC")
    thirty_days_ago = pd.to_datetime('today', utc=True, format='mixed', dayfirst=False) - timedelta(days=30)
    seven_days_ago = pd.to_datetime('today', utc=True, format='mixed', dayfirst=False) - timedelta(days=7)
    df['Creation Time (UTC)'] = pd.to_datetime(df['Creation Time (UTC)'], utc=True, format='mixed', dayfirst=False)
    received_df = df[df['Direction'] == 'Inbound'][['From']]
    sent_df = df[(df['Direction'] == 'Outbound') & (df['Creation Time (UTC)'] >= thirty_days_ago)][['To']]
    sent_seven_df = df[(df['Direction'] == 'Outbound') & (df['Creation Time (UTC)'] >= seven_days_ago)][['To']]
    received_df['From'] = received_df['From'].apply(
        lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) else x
    ).str[1:]

    sent_df['To'] = sent_df['To'].apply(
        lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) else x
    ).str[1:]

    sent_seven_df['To'] = sent_seven_df['To'].apply(
        lambda x: ''.join(filter(str.isdigit, str(x))) if pd.notna(x) else x
    ).str[1:]


    received_df.drop_duplicates(subset=['From'], inplace=True)
    sent_df.drop_duplicates(subset=['To'], inplace=True)
    sent_seven_df.drop_duplicates(subset=['To'], inplace=True)

    # get_update_df(received_df, 'CCM+CH+MVPC+MVPT+JC+RC+PD')
    global mvp_df
    received_df.columns = [0]
    mvp_df = pd.concat([mvp_df, received_df], ignore_index=True)
    received_df = None
    # get_update_df(sent_df, 'RCSMS-Sent')

    # with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     received_df.to_excel(writer, sheet_name='RCSMS-Received', index=False, header=False)

    # with pd.ExcelWriter('./data/List Cleaner.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    #     sent_df.to_excel(writer, sheet_name='RCSMS-Sent', index=False, header=False)

    outbound_df_list.append(sent_df.values.tolist())
    sent_df = None

    outbound_seven_days_list.append(sent_seven_df.values.tolist())
    sent_seven_df = None

def add_c3(df: pd.DataFrame, dbx):
    print("Processing C3")
    global dnc_df
    df['Contact Information'] = df['Contact Information'].astype(str)
    df['Contact Information'] = df['Contact Information'].str.replace(r'\D', '', regex=True)
    output_df = df[df['Contact Information'].str.len() == 10][['Contact Information']].drop_duplicates(subset=['Contact Information'])
    # get_update_df(output_df, 'DNC')
    output_df.columns = [0]
    dnc_df = pd.concat([dnc_df, output_df], ignore_index=True)
    output_df = None

def get_latest_file(folder_path, dbx):
    try:
        # List files in the folder
        response = dbx.files_list_folder(folder_path)

        latest_file = None
        latest_timestamp = None

        for entry in response.entries:
            if isinstance(entry, dropbox.files.FileMetadata):
                # Compare the 'client_modified' timestamp to find the latest file
                if latest_timestamp is None or entry.client_modified > latest_timestamp:
                    latest_file = entry
                    latest_timestamp = entry.client_modified

        return latest_file.path_lower if latest_file else None

    except Exception as e:
        print(f"Error getting latest file from {folder_path}: {e}")
        return None

def load_sheets(root_path, dbx, conversion_dict):
    try:
        # List files and folders in the current path
        response = dbx.files_list_folder(root_path)
        for entry in response.entries:
            if isinstance(entry, dropbox.files.FileMetadata):
                file_path = entry.path_lower
                folder_path = "/".join(file_path.split('/')[:-1])
                folder_name = file_path.split('/')[:-1][-1]

                # Check if valid folder name then get corresponding function per folder
                if folder_name in conversion_dict:
                    file_type_function = conversion_dict[folder_name]

                    # Process the latest only
                    if file_path == get_latest_file(folder_path, dbx):
                        file_type_function(file_path, dbx)
                        
            elif isinstance(entry, dropbox.files.FolderMetadata):
                load_sheets(entry.path_lower, dbx, conversion_dict)

        # Handle pagination
        while response.has_more:
            response = dbx.files_list_folder_continue(response.cursor)
            for entry in response.entries:
                if isinstance(entry, dropbox.files.FileMetadata):
                    file_path = entry.path_lower
                    folder_path = "/".join(file_path.split('/')[:-1])
                    folder_name = file_path.split('/')[:-1][-1]

                    # Check if valid folder name then get corresponding function per folder
                    if folder_name in conversion_dict:
                        file_type_function = conversion_dict[folder_name]

                        # Process the latest only
                        if file_path == get_latest_file(folder_path, dbx):
                            file_type_function(file_path, dbx)
                            
                elif isinstance(entry, dropbox.files.FolderMetadata):
                    load_sheets(entry.path_lower, dbx, conversion_dict)

    except dropbox.exceptions.ApiError as e:
        print(f"Error accessing path '{root_path}': {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

def concat_contact_center_files(path: str, dbx):

    result = dbx.files_list_folder(path)
    contact_center_files = result.entries
    df_list = []

    for file in contact_center_files:
        if isinstance(file, dropbox.files.FileMetadata):
            file_path = file.path_lower
            df = read_dropbox_file(file_path, dbx)
            df_list.append(df)
    
    if df_list:
        combined_df = pd.concat(df_list)
        return combined_df
    
def concat_inbound_contact_history(path: str, dbx: dropbox.Dropbox):

    result = dbx.files_list_folder(path)
    inbound_contact_history_files = result.entries
    df_list = []

    for file in inbound_contact_history_files:
        if isinstance(file, dropbox.files.FileMetadata):
            file_path = file.path_lower
            df = read_dropbox_file(file_path, dbx)
            df_list.append(df)
    
    if df_list:
        combined_df = pd.concat(df_list)
        return combined_df

    
def concat_rc_files(path: str, dbx):

    result = dbx.files_list_folder(path)
    rc_files = result.entries
    df_list = []

    for file in rc_files:
        if isinstance(file, dropbox.files.FileMetadata):
            file_path = file.path_lower
            df = read_dropbox_file(file_path, dbx)
            df_list.append(df)
    
    if df_list:
        combined_df = pd.concat(df_list)
        return combined_df

def concat_mvp_files(path: str, dbx):

    result = dbx.files_list_folder(path)
    mvp_files = result.entries
    df_list = []

    for file in mvp_files:
        if isinstance(file, dropbox.files.FileMetadata):
            file_path = file.path_lower
            df = read_dropbox_file(file_path, dbx)
            df_list.append(df)
    
    if df_list:
        combined_df = pd.concat(df_list)
        return combined_df
    
def concat_inbound_mvp_calls(path: str, dbx: dropbox.Dropbox):

    result = dbx.files_list_folder(path)
    inbound_mvp_call_files = result.entries
    df_list = []

    for file in inbound_mvp_call_files:
        if isinstance(file, dropbox.files.FileMetadata):
            file_path = file.path_lower
            if file_path.endswith('.xlsx'):
                metadata, response = dbx.files_download(file_path)
                df = pd.read_excel(BytesIO(response.content), sheet_name='Calls')
                df_list.append(df)
    
    if df_list:
        combined_df = pd.concat(df_list)
        return combined_df

def concat_c3(path: str, dbx):

    result = dbx.files_list_folder(path)
    mvp_files = result.entries
    df_list = []

    for file in mvp_files:
        if isinstance(file, dropbox.files.FileMetadata):
            file_path = file.path_lower
            df = read_dropbox_file(file_path, dbx)
            df_list.append(df)
    
    if df_list:
        combined_df = pd.concat(df_list)
        return combined_df

def create_local_list_cleaner(local_data_path: str, root_path: str, dbx: dropbox.Dropbox) -> None:
    global dnc_df, mvp_df, db_id_df, time_df, conv_df, pd_jr_aa_df

    print("Creating new local list cleaner file")

    sheet_names = [
        "CCM+CH+MVPC+MVPT+JC+RC+PD (Cold)",
        "DNC (Cold-PD)",
        "UniqueDB ID (Cold)",
        "CallOut-14d+TextOut-30d (Cold)",
        "PDConvDup (PD)",
        "PDJRAADups (PD)"
    ]

    for sheet_name in sheet_names:
        dropbox_path = f"{root_path}/{sheet_name}.csv"
        local_file_path = f"{local_data_path}/{sheet_name}.csv"
        metadata, response = dbx.files_download(dropbox_path)

        with open(local_file_path, 'wb') as f:
            f.write(response.content)

        df = pd.read_csv(local_file_path, low_memory=False, header=None)

        match sheet_name:
            case "CCM+CH+MVPC+MVPT+JC+RC+PD (Cold)":
                mvp_df = df
            case "DNC (Cold-PD)":
                dnc_df = df
            case "UniqueDB ID (Cold)":
                db_id_df = df
            case "CallOut-14d+TextOut-30d (Cold)":
                time_df = df
            case "PDConvDup (PD)":
                conv_df = df
            case "PDJRAADups (PD)":
                pd_jr_aa_df = df
            case _:
                print(f"Unknown sheet: {sheet_name}")

def check_user_folder_paths(dbx: dropbox.Dropbox):
    result = dbx.files_list_folder(path="", recursive=True)
    for entry in result.entries:
        if isinstance(entry, dropbox.files.FolderMetadata):
            if entry.path_display == '/List Cleaner & JC DNC':
                return entry.path_display
            
def update_latest_cleaner_file_label(app_window: ctk.CTkFrame, dbx: dropbox.Dropbox):
    metadata = dbx.files_get_metadata('/List Cleaner & JC DNC/DNC (Cold-PD).csv')
    last_modified_date = metadata.client_modified
    app_window.last_update_label.configure(text=f'List cleaner file last update: {last_modified_date}')

def drop_list_cleaner_dupes(file_path: str):
    print("Removing duplicates")
    sheets = pd.ExcelFile(file_path).sheet_names
    dataframes = {}
    
    for sheet in sheets:
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)
        dataframes[sheet] = df.drop_duplicates().dropna()

    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=None)

def main(auth_code: str, app_window: ctk.CTkFrame):

    try:
        dbx = dropbox.Dropbox(auth_code)

        root_path = check_user_folder_paths(dbx)
        global outbound_df_list, outbound_seven_days_list
        outbound_df_list = []
        outbound_seven_days_list = []
        
        # Create constant variables
        local_data_path = './data'
        conversion_dict = {
            "jc": add_jc,
            "pd_db": add_unique_db,
            "pd_phone": add_pd_phones,
            "pd_remove": add_remove_list,
            "pd_convdups": add_pd_conv_dup,
            "sly": add_sly,
            "pd_jr_aa": add_pd_jr_aa_dup
        }

        # Download and replace local list cleaner file
        create_local_list_cleaner(local_data_path, root_path, dbx)

        # Update the list cleaner file
        load_sheets(root_path, dbx, conversion_dict)

        # Process c3 folder files
        c3_df = concat_c3(f'{root_path}/c3', dbx)
        add_c3(c3_df, dbx)

        # Process contact_center folder files
        contact_center_df = concat_contact_center_files(f'{root_path}/contact_center', dbx)
        add_contact_center(contact_center_df, dbx)

        # Process contact_history folder files
        contact_history_df = concat_inbound_contact_history(f'{root_path}/contact_history', dbx)
        add_contact_history_inbound(contact_history_df, dbx)

        # Process rc folder files
        rc_df = concat_rc_files(f'{root_path}/rc', dbx)
        add_rc(rc_df, dbx)

        # Process mvp folder files
        mvp_df = concat_mvp_files(f'{root_path}/mvp', dbx)
        add_mvp(mvp_df, dbx)

        # Process mvp_calls folder files
        mvp_calls_df = concat_inbound_mvp_calls(f'{root_path}/mvp_calls', dbx)
        add_mvp_calls_inbound(mvp_calls_df, dbx)

        # # Save all new sheets locally
        # append_to_multiple_sheets(local_list_cleaner_path)

        # Create and replace the outbound sheet
        save_all_files()

        # # Drop all duplicates of list cleaner file
        # drop_list_cleaner_dupes(local_list_cleaner_path)

        # Upload to dropbox
        export_to_dropbox(root_path, local_data_path, dbx)

        # Update label
        update_latest_cleaner_file_label(app_window, dbx)
    
    except Exception as e:
        print(f"An error occured: {e}")
        raise RuntimeError

if __name__ == "__main__":
    main()
